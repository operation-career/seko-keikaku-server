from flask import Flask, request, jsonify, send_file
import io
import openpyxl
import requests

app = Flask(__name__)

TEMPLATE_FILE = 'template.xlsx'  # 本物の施工計画書テンプレートファイル名

@app.route('/webhook', methods=['POST'])
def webhook():
    try:
        data = request.get_json()

        file_url = data.get('file_url')
        vehicle_equipment = data.get('使用車両備品')
        work_description = data.get('主な作業内容')

        # ファイル（PDF）ダウンロード（ここでは未使用、後で使う想定）
        file_response = requests.get(file_url)
        if file_response.status_code != 200:
            raise Exception("PDFファイルのダウンロードに失敗しました。")

        # 仮データ（あとでPDF解析結果と入れ替え）
        工事名 = "仮 工事名"
        工事場所 = "仮 工事場所"
        工期 = "仮 工期"

        # Excelファイルを作成
        output = create_excel(工事名, 工事場所, 工期, vehicle_equipment, work_description)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name="施工計画書.xlsx")

    except Exception as e:
        print(f"エラー: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

def create_excel(工事名, 工事場所, 工期, vehicle_equipment, work_description):
    # 本物のテンプレートファイルを開く
    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb.active  # 1番目のシートを対象とする

    # 指定のセルに書き込む（例：適宜あなたのテンプレートに合わせて変更）
    ws['B2'] = 工事名         # 工事名をB2セルに
    ws['B3'] = 工事場所       # 工事場所をB3セルに
    ws['B4'] = 工期           # 工期をB4セルに
    ws['B6'] = vehicle_equipment  # 使用車両・備品（仮：B6に）
    ws['B7'] = work_description   # 主な作業内容（仮：B7に）

    # メモリ上に保存
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    return file

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
